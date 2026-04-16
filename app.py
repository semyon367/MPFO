import streamlit as st
import io
import re
from collections import defaultdict
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

# --- КОНСТАНТЫ ---
SHEET_NAME = "Детализация МП Инспектор"

COLUMN_KEYWORDS = {
    "subjekt": ["субъект рф"],
    "podrazd": ["подразделение"],
    "vid_nadzora": ["вид надзора"],
    "nom_knm": ["номер кнм"],
    "vid": ["вид"],
    "status": ["статус кнм"],
    "narusheniya": ["нарушения выявлены"],
    "proverka_ogv": ["проверка огв/омсу"],
    "knd": ["кнд"],
    "ssylki": ["ссылки на файлы"],
    "date_act": ["дата составления акта о результате кнм", "дата составления акта"],
    "s_vks": ["с вкс", "вкс"],
}

DISTRICTS = [
    (
        "ЦФО",
        "Центральный ФО",
        [
            "УНДиПР ГУ МЧС России по Тверской области",
            "УНДиПР ГУ МЧС России по Курской области",
            "УНДиПР ГУ МЧС России по г. Москве",
            "УНДиПР ГУ МЧС России по Московской области",
            "УНДиПР ГУ МЧС России по Владимирской области",
            "УНДиПР ГУ МЧС России по Тамбовской области",
            "УНДиПР ГУ МЧС России по Тульской области",
            "УНДиПР ГУ МЧС России по Липецкой области",
            "УНДиПР ГУ МЧС России по Рязанской области",
            "УНДиПР ГУ МЧС России по Костромской области",
            "УНДиПР ГУ МЧС России по Ярославской области",
            "УНДиПР ГУ МЧС России по Ивановской области",
            "УНДиПР ГУ МЧС России по Воронежской области",
            "УНДиПР ГУ МЧС России по Калужской области",
            "УНДиПР ГУ МЧС России по Белгородской области",
            "УНДиПР ГУ МЧС России по Брянской области",
            "УНДиПР ГУ МЧС России по Смоленской области",
            "УНДПР ГУ МЧС России по Орловской области",
        ],
    ),
    (
        "СЗФО",
        "Северо-Западный ФО",
        [
            "УНДПР Главного управления МЧС России по г. Санкт-Петербургу",
            "УНДиПР ГУ МЧС России по Ленинградской области",
            "УНДиПР ГУ МЧС России по Калининградской области",
            "УНДиПР ГУ МЧС России по Псковской области",
            "УНДиПР ГУ МЧС России по Республике Коми",
            "УНДиПР ГУ МЧС России по Архангельской области",
            "УНДиПР ГУ МЧС России по Вологодской области",
            "УНДиПР ГУ МЧС России по Новгородской области",
            "УНДиПР ГУ МЧС России по Республике Карелия",
            "УНДиПР ГУ МЧС России по Мурманской области",
            "ОНДиПР ГУ МЧС России по Ненецкому автономному округу",
        ],
    ),
    (
        "СКФО",
        "Северо-Кавказский ФО",
        [
            "УНДиПР ГУ МЧС России по Кабардино-Балкарской Республике",
            "УНДиПР ГУ МЧС России по Республике Северная Осетия - Алания",
            "УНДиПР ГУ МЧС России по Республике Дагестан",
            "УНДиПР ГУ МЧС России по Карачаево-Черкесской Республике",
            "УНДиПР ГУ МЧС России по Ставропольскому краю",
            "УНДиПР ГУ МЧС России по Республике Ингушетия",
            "УНДиПР ГУ МЧС России по Чеченской Республике",
        ],
    ),
    (
        "ЮФО",
        "Южный ФО",
        [
            "УНДиПР ГУ МЧС России по г. Севастополю",
            "УНДиПР ГУ МЧС России по Волгоградской области",
            "УНДиПР ГУ МЧС России по Ростовской области",
            "УНДиПР ГУ МЧС России по Республике Адыгея",
            "УНДиПР ГУ МЧС России по Астраханской области",
            "УНДиПР ГУ МЧС России по Республике Крым",
            "УНДиПР ГУ МЧС России по Республике Калмыкия",
            "УНДиПР ГУ МЧС России по Краснодарскому краю",
        ],
    ),
    (
        "ПФО",
        "Приволжский ФО",
        [
            "УНДиПР ГУ МЧС России по Пензенской области",
            "УНДиПР ГУ МЧС России по Оренбургской области",
            "УНДиПР ГУ МЧС России по Ульяновской области",
            "УНДиПР ГУ МЧС России по Республике Башкортостан",
            "УНДиПР ГУ МЧС России по Удмуртской Республике",
            "УНДиПР ГУ МЧС России по Самарской области",
            "УНДиПР ГУ МЧС России по Нижегородской области",
            "УНДиПР ГУ МЧС России по Кировской области",
            "УНДиПР ГУ МЧС России по Пермскому краю",
            "УНДиПР ГУ МЧС России по Саратовской области",
            "УНДиПР ГУ МЧС России по Республике Мордовия",
            "УНДиПР ГУ МЧС России по Чувашской Республике - Чувашии",
            "УНДиПР Главного управления МЧС России по Республике Марий Эл",
            "УНДиПР ГУ МЧС России по Республике Татарстан",
        ],
    ),
    (
        "ДФО",
        "Дальневосточный ФО",
        [
            "УНДиПР ГУ МЧС России по Чукотскому АО",
            "УНДиПР ГУ МЧС России по Забайкальскому краю",
            "УНДиПР ГУ МЧС России по Сахалинской области",
            "УНДиПР ГУ МЧС России по Приморскому краю",
            "УНДиПР ГУ МЧС России по Еврейской АО",
            "УНДиПР ГУ МЧС России по Камчатскому краю",
            "УНДиПР ГУ МЧС России по Республике Саха (Якутия)",
            "УНДиПР Главного управления МЧС России по Республике Бурятия",
            "УНДиПР ГУ МЧС России по Хабаровскому краю",
            "УНДиПР ГУ МЧС России по Магаданской области",
            "УНДиПР ГУ МЧС России по Амурской области",
        ],
    ),
    (
        "СФО",
        "Сибирский ФО",
        [
            "УНДиПР ГУ МЧС России по Республике Тыва",
            "УНДиПР ГУ МЧС России по Новосибирской области",
            "УНДиПР ГУ МЧС России по Кемеровской области - Кузбассу",
            "УНДиПР ГУ МЧС России по Красноярскому краю",
            "УНДиПР ГУ МЧС России по Томской области",
            "УНДиПР ГУ МЧС России по Республике Алтай",
            "УНДиПР ГУ МЧС России по Омской области",
            "УНДиПР ГУ МЧС России по Республике Хакасия",
            "УНДиПР ГУ МЧС России по Алтайскому краю",
            "УНДиПР ГУ МЧС России по Иркутской области",
        ],
    ),
    (
        "УФО",
        "Уральский ФО",
        [
            "УНДиПР ГУ МЧС России по Курганской области",
            "УНДиПР ГУ МЧС России по Свердловской области",
            "УНДиПР ГУ МЧС России по Челябинской области",
            "УНДиПР ГУ МЧС России по Ямало-Ненецкому АО",
            "УНДиПР ГУ МЧС России по Ханты-Мансийскому АО - Югре",
            "УНДиПР ГУ МЧС России по Тюменской области",
        ],
    ),
]

NEW_REGIONS = [
    "УНДиПР ГУ МЧС России по Донецкой Народной Республике",
    "УНДиПР ГУ МЧС России по Запорожской области",
    "УНДиПР ГУ МЧС России по Херсонской области",
    "УНДиПР ГУ МЧС России по Луганской Народной Республике",
]

ALL_SUBJECTS = {subject for _, _, subjects in DISTRICTS for subject in subjects}
ALL_SUBJECTS.update(NEW_REGIONS)
CANONICAL_SUBJECTS = {normalize_key: subject for subject in ALL_SUBJECTS for normalize_key in [re.sub(r"\s+", " ", subject.strip()).lower()]}


# --- ФУНКЦИИ ОБРАБОТКИ ДАННЫХ ---

def normalize_str(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def find_column_index(headers, possible_names):
    headers_norm = [normalize_str(header) for header in headers]
    possible_norm = [normalize_str(name) for name in possible_names]

    for idx, norm in enumerate(headers_norm):
        if norm in possible_norm:
            return idx

    for idx, norm in enumerate(headers_norm):
        for possible_name in possible_norm:
            if possible_name in norm:
                return idx

    return None


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


@st.cache_data
def load_data(uploaded_file):
    if uploaded_file is None:
        return None, None

    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    except Exception as e:
        st.error(f"Ошибка при открытии файла: {e}")
        return None, None

    if SHEET_NAME not in wb.sheetnames:
        st.error(f"Лист '{SHEET_NAME}' не найден в файле.")
        return None, None

    ws = wb[SHEET_NAME]
    headers = [cell.value if cell.value else "" for cell in ws[1]]

    col_indices = {}
    missing = []

    for key, possible_names in COLUMN_KEYWORDS.items():
        idx = find_column_index(headers, possible_names)
        if idx is None:
            if key == "podrazd" and len(headers) > 17:
                idx = 17
            else:
                missing.append(f"'{key}' (искали: {possible_names})")
        col_indices[key] = idx

    if missing:
        st.warning(f"Не найдены столбцы: {', '.join(missing)}. Некоторые функции могут не работать.")
        # Не прерываем выполнение, чтобы попробовать обработать то, что есть

    data = [
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if not all(cell is None for cell in row)
    ]
    
    return data, col_indices


def filter_by_date(data, col_idx, date_from, date_to):
    if "date_act" not in col_idx or col_idx["date_act"] is None:
        st.error("Столбец с датой не найден.")
        return [], 0, 0

    date_col = col_idx["date_act"]
    filtered = []
    skipped_out_of_range = 0
    skipped_invalid_date = 0

    for row in data:
        # Проверка границ индекса
        if date_col >= len(row):
            skipped_invalid_date += 1
            continue
            
        parsed = parse_date(row[date_col])
        if parsed is None:
            skipped_invalid_date += 1
            continue
        if date_from <= parsed <= date_to:
            filtered.append(row)
        else:
            skipped_out_of_range += 1

    return filtered, skipped_out_of_range, skipped_invalid_date


def calculate_metrics_by_subject(data, col_idx):
    # Проверка наличия всех необходимых колонок
    required_keys = ["subjekt", "podrazd", "nom_knm", "vid", "status", "proverka_ogv", "vid_nadzora", "knd", "narusheniya", "s_vks", "ssylki"]
    if not all(k in col_idx and col_idx[k] is not None for k in required_keys):
        return {}

    subj_col = col_idx["subjekt"]
    podrazd_col = col_idx["podrazd"]
    knm_col = col_idx["nom_knm"]
    vid_col = col_idx["vid"]
    status_col = col_idx["status"]
    proverka_col = col_idx["proverka_ogv"]
    vid_nadzora_col = col_idx["vid_nadzora"]
    knd_col = col_idx["knd"]
    nar_col = col_idx["narusheniya"]
    vks_col = col_idx["s_vks"]
    ssylki_col = col_idx["ssylki"]

    allowed_vids = {"", "выездная проверка", "рейдовый осмотр", "инспекционный визит"}
    knm_info = {}

    for row in data:
        # Проверка длины строки
        max_idx = max([subj_col, knm_col, status_col, proverka_col, vid_nadzora_col, vid_col, knd_col, nar_col, vks_col, ssylki_col])
        if len(row) <= max_idx:
            continue

        reasons_base = []
        if normalize_str(row[status_col]) != "завершена":
            reasons_base.append("status")
        if normalize_str(row[proverka_col]) != "нет":
            reasons_base.append("proverka")
        if normalize_str(row[vid_nadzora_col]) == "гнго":
            reasons_base.append("vid_nadzora")

        subject_raw = str(row[subj_col]).strip() if row[subj_col] else ""
        subject_name = CANONICAL_SUBJECTS.get(normalize_str(subject_raw), "")
        knm = row[knm_col]
        if not row[subj_col] or not knm:
            reasons_base.append("empty")

        if reasons_base or subject_name not in ALL_SUBJECTS:
            continue

        vid_val = normalize_str(row[vid_col]) if row[vid_col] else ""
        knd_str = normalize_str(row[knd_col]) if row[knd_col] else ""
        nar_str = normalize_str(row[nar_col]) if row[nar_col] else ""
        vks_str = normalize_str(row[vks_col]) if row[vks_col] else ""
        ssylki_ok = row[ssylki_col] is not None and str(row[ssylki_col]).strip() != ""

        if knm not in knm_info:
            knm_info[knm] = {
                "subject": subject_name,
                "vks_denom": False,
                "vks_num": False,
                "och_denom": False,
                "och_num": False,
                "och_nar": False,
            }

        info = knm_info[knm]
        if vid_val in allowed_vids:
            info["vks_denom"] = True
            if vks_str == "да" and ssylki_ok:
                info["vks_num"] = True

        if vid_val in allowed_vids and "осмотр" in knd_str:
            info["och_denom"] = True
            if vks_str == "нет" and ssylki_ok:
                info["och_num"] = True
            if nar_str == "да":
                info["och_nar"] = True

    metrics = defaultdict(lambda: [set(), set(), set(), set(), set()])

    for knm, info in knm_info.items():
        subject = info["subject"]
        if info["vks_denom"]:
            metrics[subject][0].add(knm)
        if info["vks_num"]:
            metrics[subject][1].add(knm)
        if info["och_denom"]:
            metrics[subject][2].add(knm)
        if info["och_num"]:
            metrics[subject][3].add(knm)
        if info["och_nar"]:
            metrics[subject][4].add(knm)

    return metrics


def make_subject_rows(subjects, metrics_year, metrics_week):
    rows = []
    for subject in subjects:
        year_sets = metrics_year.get(subject, [set(), set(), set(), set(), set()])
        week_sets = metrics_week.get(subject, [set(), set(), set(), set(), set()])
        rows.append(
            {
                "subject": subject,
                "vks_total_year": len(year_sets[0]),
                "vks_mp_year": len(year_sets[1]),
                "vks_total_week": len(week_sets[0]),
                "vks_mp_week": len(week_sets[1]),
                "och_total_year": len(year_sets[2]),
                "och_mp_year": len(year_sets[3]),
                "och_nar_year": len(year_sets[4]),
                "och_total_week": len(week_sets[2]),
                "och_mp_week": len(week_sets[3]),
                "och_nar_week": len(week_sets[4]),
            }
        )
    return rows


def fmt_ratio(num, den):
    pct = num / den * 100 if den > 0 else 0.0
    return f"{num} ({pct:.2f}%)"


def fmt_och(total, nar):
    return f"{total} ({nar})"


def auto_adjust_row_heights(ws, start_row, end_row):
    col_a_width = ws.column_dimensions["A"].width or 60
    
    for row_idx in range(start_row, end_row + 1):
        max_lines = 1
        cell = ws.cell(row_idx, 1)
        
        if cell.value:
            text = str(cell.value)
            chars_per_line = int(col_a_width * 1.2)
            lines = len(text) / chars_per_line if chars_per_line > 0 else 1
            max_lines = max(1, int(lines) + 1)
        
        row_height = max_lines * 15
        row_height = min(row_height, 60)
        row_height = max(row_height, 20)
        
        ws.row_dimensions[row_idx].height = row_height


def generate_report(district_rows, selected_date, week_start):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    blue_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    green_fill = PatternFill("solid", start_color="00FF00", end_color="00FF00")
    yellow_fill = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
    red_fill = PatternFill("solid", start_color="FF0000", end_color="FF0000")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row_num):
        for cell in ws[row_num]:
            cell.fill = blue_fill
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    for sheet_name, district_name, rows in district_rows:
        ws = wb.create_sheet(title=sheet_name)

        ws.append(["Субъект", "ВКС", "", "", "", "ОЧНЫЕ", "", "", ""])
        ws.merge_cells("A1:A2")
        ws.merge_cells("B1:E1")
        ws.merge_cells("F1:I1")

        ws.append(
            [
                "",
                "Всего в ААС КНД с начала года",
                "Всего в МП Инспектор с начала года",
                "Всего в ААС КНД за прошедшую неделю",
                "Всего в МП Инспектор за прошедшую неделю",
                "Всего в ААС КНД с начала года (из них с нарушениями)",
                "Всего в МП Инспектор с начала года",
                "Всего в ААС КНД за прошедшую неделю (из них с нарушениями)",
                "Всего в МП Инспектор за прошедшую неделю",
            ]
        )

        style_header_row(ws, 1)
        style_header_row(ws, 2)
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 60

        ws.append([district_name, "", "", "", "", "", "", "", ""])
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)
        district_cell = ws.cell(3, 1)
        district_cell.font = Font(bold=True, name="Arial", size=10)
        district_cell.alignment = Alignment(horizontal="center", vertical="center")

        def get_och_ratio(row):
            if row["och_total_year"] > 0:
                return row["och_mp_year"] / row["och_total_year"]
            return 0.0
        
        sorted_rows = sorted(rows, key=get_och_ratio)

        start_row = 4
        for row_idx, row_data in enumerate(sorted_rows, start=start_row):
            ws.append(
                [
                    row_data["subject"],
                    row_data["vks_total_year"],
                    fmt_ratio(row_data["vks_mp_year"], row_data["vks_total_year"]),
                    row_data["vks_total_week"],
                    fmt_ratio(row_data["vks_mp_week"], row_data["vks_total_week"]),
                    fmt_och(row_data["och_total_year"], row_data["och_nar_year"]),
                    fmt_ratio(row_data["och_mp_year"], row_data["och_total_year"]),
                    fmt_och(row_data["och_total_week"], row_data["och_nar_week"]),
                    fmt_ratio(row_data["och_mp_week"], row_data["och_total_week"]),
                ]
            )

            # --- Color Logic ---
            vks_total_y = row_data["vks_total_year"]
            vks_mp_y = row_data["vks_mp_year"]
            och_total_y = row_data["och_total_year"]
            och_mp_y = row_data["och_mp_year"]

            vks_total_w = row_data["vks_total_week"]
            vks_mp_w = row_data["vks_mp_week"]
            och_total_w = row_data["och_total_week"]
            och_mp_w = row_data["och_mp_week"]

            vks_ratio_y = (vks_mp_y / vks_total_y) if vks_total_y > 0 else 0.0
            och_ratio_y = (och_mp_y / och_total_y) if och_total_y > 0 else 0.0
            vks_ratio_w = (vks_mp_w / vks_total_w) if vks_total_w > 0 else 0.0
            och_ratio_w = (och_mp_w / och_total_w) if och_total_w > 0 else 0.0

            # Условия по годам
            year_vks_ok = vks_ratio_y > 0.10
            year_och_ok = och_ratio_y > 0.80
            year_both_ok = year_vks_ok and year_och_ok
            
            # Условия по неделе
            week_vks_ok = vks_ratio_w >= 0.10
            week_och_ok = och_ratio_w >= 0.80
            week_both_ok = week_vks_ok and week_och_ok
            
            # --- Зелёный: столбец A (субъект) ---
            cell_a = ws.cell(row_idx, 1)
            cell_a.border = border
            cell_a.font = Font(name="Arial", size=10)
            cell_a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if year_both_ok:
                cell_a.fill = green_fill
            
            # --- Жёлтый: столбцы C и G (с начала года) ---
            cell_c = ws.cell(row_idx, 3)
            cell_c.border = border
            cell_c.font = Font(name="Arial", size=10)
            cell_c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if not year_vks_ok:
                cell_c.fill = yellow_fill
            
            cell_g = ws.cell(row_idx, 7)
            cell_g.border = border
            cell_g.font = Font(name="Arial", size=10)
            cell_g.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if not year_och_ok:
                cell_g.fill = yellow_fill
            
            # --- Красный: столбцы E и I (за неделю) ---
            if not year_both_ok and not week_both_ok:
                cell_e = ws.cell(row_idx, 5)
                cell_e.border = border
                cell_e.font = Font(name="Arial", size=10)
                cell_e.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if not week_vks_ok:
                    cell_e.fill = red_fill
                
                cell_i = ws.cell(row_idx, 9)
                cell_i.border = border
                cell_i.font = Font(name="Arial", size=10)
                cell_i.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if not week_och_ok:
                    cell_i.fill = red_fill

        total_vks_year = sum(row["vks_total_year"] for row in rows)
        total_vks_mp_year = sum(row["vks_mp_year"] for row in rows)
        total_vks_week = sum(row["vks_total_week"] for row in rows)
        total_vks_mp_week = sum(row["vks_mp_week"] for row in rows)
        total_och_year = sum(row["och_total_year"] for row in rows)
        total_och_mp_year = sum(row["och_mp_year"] for row in rows)
        total_och_nar_year = sum(row["och_nar_year"] for row in rows)
        total_och_week = sum(row["och_total_week"] for row in rows)
        total_och_mp_week = sum(row["och_mp_week"] for row in rows)
        total_och_nar_week = sum(row["och_nar_week"] for row in rows)

        ws.append(
            [
                f"Итого за {district_name}",
                total_vks_year,
                fmt_ratio(total_vks_mp_year, total_vks_year),
                total_vks_week,
                fmt_ratio(total_vks_mp_week, total_vks_week),
                fmt_och(total_och_year, total_och_nar_year),
                fmt_ratio(total_och_mp_year, total_och_year),
                fmt_och(total_och_week, total_och_nar_week),
                fmt_ratio(total_och_mp_week, total_och_week),
            ]
        )
        total_row_idx = ws.max_row

        ws.column_dimensions["A"].width = 60.89
        ws.column_dimensions["B"].width = 9.33
        ws.column_dimensions["C"].width = 13.22
        ws.column_dimensions["D"].width = 8.89
        ws.column_dimensions["E"].width = 13.22
        ws.column_dimensions["F"].width = 13.44
        ws.column_dimensions["G"].width = 16.44
        ws.column_dimensions["H"].width = 9.44
        ws.column_dimensions["I"].width = 14.33

        auto_adjust_row_heights(ws, start_row, total_row_idx - 1)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=9):
            for cell in row:
                if isinstance(cell, openpyxl.cell.MergedCell):
                    continue
                cell.border = border
                if cell.row >= 4 and cell.row < total_row_idx:
                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif cell.row == total_row_idx:
                    cell.font = Font(bold=True, name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if cell.row >= 4 and not cell.font.name:
                    cell.font = Font(name="Arial", size=10)

        info_row = ws.max_row + 2
        ws.cell(info_row, 1, f"Дата отчёта: {selected_date.strftime('%d.%m.%Y')}")
        ws.cell(
            info_row + 1,
            1,
            f"Прошедшая неделя: {week_start.strftime('%d.%m.%Y')} - {selected_date.strftime('%d.%m.%Y')}",
        )

    # Сохраняем в буфер памяти
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- ИНТЕРФЕЙС STREAMLIT ---

def main():
    st.set_page_config(page_title="Анализ МП Инспектор", layout="wide")
    st.title("📊 Анализ применения МП Инспектор по федеральным округам")

    # 1. Загрузка файла
    uploaded_file = st.file_uploader("Загрузите файл выгрузки (.xlsx)", type=["xlsx", "xlsm"])
    
    if uploaded_file is not None:
        st.success("Файл загружен!")
        with st.spinner("Обработка данных..."):
            data, col_idx = load_data(uploaded_file)
            
            if data is None:
                st.stop()

            st.write(f"✅ Загружено строк: {len(data)}")

            # 2. Выбор даты
            col1, col2 = st.columns(2)
            with col1:
                selected_date = st.date_input("Дата отчёта:", value=datetime.now().date())
            
            year_start = date(selected_date.year, 1, 1)
            week_start = selected_date - timedelta(days=6)

            with col2:
                st.info(f"Период с начала года: {year_start} - {selected_date}")
                st.info(f"Период за неделю: {week_start} - {selected_date}")

            # 3. Фильтрация и расчет
            ytd_data, ytd_out, ytd_invalid = filter_by_date(data, col_idx, year_start, selected_date)
            week_data, week_out, week_invalid = filter_by_date(data, col_idx, week_start, selected_date)

            metrics_year = calculate_metrics_by_subject(ytd_data, col_idx)
            metrics_week = calculate_metrics_by_subject(week_data, col_idx)

            district_rows = []
            for short_name, full_name, subjects in DISTRICTS:
                rows = make_subject_rows(subjects, metrics_year, metrics_week)
                district_rows.append((short_name, full_name, rows))

            # 4. Генерация и скачивание
            try:
                output_buffer = generate_report(district_rows, selected_date, week_start)
                filename = f"итоги_фо_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                st.download_button(
                    label="📥 Скачать итоговый отчёт (Excel)",
                    data=output_buffer,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("Отчёт сформирован! Нажмите кнопку выше для скачивания.")
            except Exception as e:
                st.error(f"Ошибка при формировании отчёта: {e}")
    else:
        st.info("Пожалуйста, загрузите файл Excel для начала работы.")

if __name__ == "__main__":
    main()
